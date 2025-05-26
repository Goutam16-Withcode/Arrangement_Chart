# FULL UPDATED CODE WITH DYNAMIC ATTENDANCE SHEET
# Enhancements:
# - Unique attendance sheet per room
# - Roll numbers assigned per seat position dynamically

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import os
import subprocess

SEAT_POSITIONS = ["F-1", "S-1", "T-1"]
HEADER_FILL = PatternFill(start_color="D1C4E9", end_color="D1C4E9", fill_type="solid")
BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

class SeatingChartApp:
    def __init__(self, master):
        self.master = master
        master.title("Seating and Attendance Sheet Generator")
        master.geometry("600x600")
        master.configure(bg="#f0f0f0")
        self.room_details_df = None
        self.students_per_bench = None
        self.roll_numbers_lists = []
        self.roll_paths = {}
        self.roll_files_selected = {}
        self.generated_file_path = None

        # Color scheme
        self.button_bg = "#FFB300"
        self.button_hover = "#FFA000"
        self.header_bg = "#FFB300"
        self.frame_bg = "#FFFFFF"

        title_label = tk.Label(master, text="Seating Chart & Attendance Generator",
                               bg=self.header_bg, fg="white", font=("Helvetica", 18, "bold"))
        title_label.pack(fill="x", pady=(10, 10))

        frame_input = tk.Frame(master, bg=self.frame_bg, bd=2, relief=tk.RIDGE)
        frame_input.pack(padx=10, pady=5, fill="x")
        input_header = tk.Label(frame_input, text="📂 Input Files", bg=self.header_bg, fg="white",
                                font=("Helvetica", 16, "bold"))
        input_header.pack(fill="x")

        lbl_room = tk.Label(frame_input, text="Room Layout File:", font=("Helvetica", 12, "bold"), bg=self.frame_bg)
        lbl_room.pack(anchor="w", pady=(10, 0), padx=10)
        self.create_button(frame_input, "📁 Upload Room Layout File", self.load_room_file)

        lbl_roll = tk.Label(frame_input, text="Roll Numbers Files:", font=("Helvetica", 12, "bold"), bg=self.frame_bg)
        lbl_roll.pack(anchor="w", pady=(10, 0), padx=10)
        for pos in ["Left", "Middle", "Right"]:
            self.roll_files_selected[pos] = False
            self.create_button(frame_input, f"🧾 Upload {pos} Roll Numbers", lambda p=pos: self.load_roll_file(p))

        frame_actions = tk.Frame(master, bg=self.frame_bg, bd=2, relief=tk.RIDGE)
        frame_actions.pack(padx=10, pady=10, fill="x")
        actions_header = tk.Label(frame_actions, text="⚙️ Actions", bg=self.header_bg, fg="white",
                                  font=("Helvetica", 16, "bold"))
        actions_header.pack(fill="x")

        self.generate_button = self.create_button(frame_actions, "✅ Generate Seating & Attendance",
                                                  self.generate_chart, active=False)
        self.download_button = self.create_button(frame_actions, "⬇️ Save File As...",
                                                  self.download_file, active=False)
        self.open_button = self.create_button(frame_actions, "📂 Open File",
                                              self.open_file, active=False)

        self.progress_var = tk.DoubleVar()
        style = ttk.Style(master)
        style.theme_use('clam')
        style.configure("TProgressbar", thickness=20, troughcolor="#EEEEEE", background=self.button_bg)
        self.progress = ttk.Progressbar(frame_actions, orient="horizontal", length=400,
                                        mode="determinate", variable=self.progress_var, style="TProgressbar")
        self.progress.pack(pady=20)

        self.status_label = tk.Label(frame_actions, text="", fg="green", bg=self.frame_bg, font=("Helvetica", 10))
        self.status_label.pack()

    def create_button(self, parent, text, command, active=True):
        btn = tk.Button(parent, text=text, command=command, width=30, pady=5,
                        bg=self.button_bg, fg="white", font=("Helvetica", 12, "bold"),
                        relief=tk.FLAT, activebackground=self.button_hover, cursor="hand2")
        if not active:
            btn.config(state=tk.DISABLED)
        btn.pack(pady=5)
        if active:
            btn.bind("<Enter>", lambda e: btn.config(bg=self.button_hover))
            btn.bind("<Leave>", lambda e: btn.config(bg=self.button_bg))
        return btn

    def load_room_file(self):
        filepath = filedialog.askopenfilename(title="Select Room Layout Excel File",
                                              filetypes=[("Excel Files", "*.xlsx *.xls")])
        if not filepath:
            return
        try:
            df = pd.read_excel(filepath)
            df.columns = df.columns.str.strip()
            required = ['Room Number', 'Number of Rows', 'Number of Bench',
                        'Number of Student per Bench', 'Left Name', 'Middle Name', 'Right Name']
            if not all(col in df.columns for col in required):
                messagebox.showerror("Error", "Room layout file is missing required columns.")
                return
            self.room_details_df = df
            self.students_per_bench = int(df['Number of Student per Bench'].iloc[0])
            self.status_label.config(text="Room layout loaded. Upload roll number files.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read room layout: {e}")

    def load_roll_file(self, position):
        filepath = filedialog.askopenfilename(title=f"Select {position} Roll Numbers File",
                                              filetypes=[("Excel Files", "*.xlsx *.xls")])
        if not filepath:
            return
        try:
            df = pd.read_excel(filepath)
            if 'Roll Number' not in df.columns:
                messagebox.showerror("Error", f"'{position}' file missing 'Roll Number' column.")
                return
            self.roll_paths[position] = df['Roll Number'].dropna().tolist()
            self.roll_files_selected[position] = True
            self.status_label.config(text=f"{position} roll numbers loaded.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read {position} roll numbers: {e}")

        if self.students_per_bench:
            needed = ["Left", "Middle", "Right"][:self.students_per_bench]
        else:
            needed = []
        if all(self.roll_files_selected.get(pos, False) for pos in needed):
            self.roll_numbers_lists = [self.roll_paths[pos] for pos in needed]
            self.generate_button.config(state=tk.NORMAL)
            self.generate_button.config(bg=self.button_bg)

    def generate_chart(self):
        try:
            wb = openpyxl.Workbook()
            for idx, row in self.room_details_df.iterrows():
                room_number = row['Room Number']
                benches = int(row['Number of Bench'])
                rows = int(row['Number of Rows'])
                total_seats = benches * rows

                att_ws = wb.create_sheet(title=f"Attendance - Room {room_number}")
                att_ws['A1'] = f"Attendance Sheet - Room {room_number}"
                att_ws['A1'].font = Font(size=14, bold=True)
                att_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
                att_ws['A2'] = "Seat Position"
                att_ws['B2'] = "Serial Number"
                att_ws['C2'] = "Roll Number"
                att_ws['D2'] = "Signature"
                for cell in ['A2', 'B2', 'C2', 'D2']:
                    att_ws[cell].font = Font(bold=True)
                    att_ws[cell].fill = HEADER_FILL
                    att_ws[cell].alignment = Alignment(horizontal='center')

                current_row = 3
                for pos_index in range(self.students_per_bench):
                    pos_label = SEAT_POSITIONS[pos_index]
                    rolls = self.roll_numbers_lists[pos_index]
                    for i, roll in enumerate(rolls[:total_seats]):
                        att_ws.cell(row=current_row, column=1, value=pos_label)
                        att_ws.cell(row=current_row, column=2, value=i+1)
                        att_ws.cell(row=current_row, column=3, value=roll)
                        att_ws.cell(row=current_row, column=4, value="")
                        for col in range(1, 5):
                            att_ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')
                            att_ws.cell(row=current_row, column=col).border = BORDER
                        current_row += 1

                self.progress_var.set((idx + 1) / len(self.room_details_df) * 100)
                self.master.update_idletasks()

            del wb['Sheet']
            output_path = os.path.join(os.getcwd(), "SeatingChart_Output.xlsx")
            wb.save(output_path)
            self.generated_file_path = output_path
            self.download_button.config(state=tk.NORMAL)
            self.open_button.config(state=tk.NORMAL)
            messagebox.showinfo("Success", f"Seating chart and attendance saved to:\n{output_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate charts: {e}")

    def download_file(self):
        if self.generated_file_path:
            save_path = filedialog.asksaveasfilename(title="Save As", defaultextension=".xlsx",
                                                     filetypes=[("Excel Files", "*.xlsx")])
            if save_path:
                with open(self.generated_file_path, 'rb') as src:
                    with open(save_path, 'wb') as dst:
                        dst.write(src.read())
                messagebox.showinfo("Saved", f"File saved to:\n{save_path}")

    def open_file(self):
        if self.generated_file_path and os.path.exists(self.generated_file_path):
            try:
                if os.name == 'nt':
                    os.startfile(self.generated_file_path)
                elif os.name == 'posix':
                    subprocess.call(['open', self.generated_file_path])
            except Exception as e:
                messagebox.showerror("Error", f"Unable to open file: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = SeatingChartApp(root)
    root.mainloop()
