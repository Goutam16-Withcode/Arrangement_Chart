import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import os
import subprocess

SEAT_POSITIONS = ["F-1", "S-1", "T-1"]
BG_COLOR = "#f0f6ff"
BTN_COLOR = "#7e57c2"
BTN_HOVER = "#9575cd"
BTN_FONT = ("Helvetica Neue", 11, "bold")
LABEL_FONT = ("Helvetica Neue", 11)
HEADER_FILL = PatternFill(start_color="D1C4E9", end_color="D1C4E9", fill_type="solid")
BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

class SeatingChartApp:
    def __init__(self, master):
        self.master = master
        master.title("🎓 Colorful Seating Chart Generator")
        master.geometry("700x750")
        master.configure(bg=BG_COLOR)

        self.room_details_df = None
        self.students_per_bench = None
        self.roll_numbers_lists = []
        self.roll_number_indices = []
        self.generated_file_path = None

        tk.Label(master, text="Seating Chart Generator", bg=BG_COLOR,
                 font=("Helvetica Neue", 16, "bold"), fg="#3f51b5").pack(pady=20)

        self.create_button("📁 Upload Room Details Excel File", self.load_room_file)
        self.status_label = tk.Label(master, text="", fg="#388e3c", bg=BG_COLOR,
                                     font=("Helvetica Neue", 10, "italic"))
        self.status_label.pack(pady=5)

        self.roll_paths = {}
        self.roll_files_selected = {pos: False for pos in ["Left", "Middle", "Right"]}
        for pos in ["Left", "Middle", "Right"]:
            self.create_button(f"🧾 Select {pos} Roll Numbers File", lambda p=pos: self.load_roll_file(p))

        self.generate_button = self.create_button("✅ Generate Seating Chart", self.generate_chart, active=False)

        self.download_label = tk.Label(master, text="", fg="#000", bg=BG_COLOR,
                                       font=("Helvetica Neue", 10, "italic"))
        self.download_label.pack(pady=10)

        self.download_button = tk.Button(master, text="⬇️ Save File As...",
                                         command=self.download_file,
                                         bg="#43a047", fg="white", font=BTN_FONT, state=tk.DISABLED)
        self.download_button.pack(pady=5)

        self.open_button = tk.Button(master, text="📂 Open Generated File",
                                     command=self.open_file,
                                     bg="#2196f3", fg="white", font=BTN_FONT, state=tk.DISABLED)
        self.open_button.pack(pady=5)

        self.progress_var = tk.DoubleVar()
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("TProgressbar", thickness=20, troughcolor='#e0e0e0', background='#4db6ac')
        self.progress = ttk.Progressbar(master, orient="horizontal", length=500,
                                        mode="determinate", variable=self.progress_var)
        self.progress.pack(pady=20)

    def create_button(self, text, command, active=True):
        btn = tk.Button(self.master, text=text, command=command,
                        bg=BTN_COLOR, fg="white", font=BTN_FONT,
                        activebackground=BTN_HOVER, activeforeground="white",
                        relief="raised", bd=2, padx=10, pady=5)
        btn.pack(pady=8)
        if not active:
            btn.config(state=tk.DISABLED)
        return btn

    def load_room_file(self):
        filepath = filedialog.askopenfilename(title="Select Room Details Excel File", filetypes=[("Excel Files", "*.xlsx *.xls")])
        if not filepath:
            return
        try:
            df = pd.read_excel(filepath)
            df.columns = df.columns.str.strip()
            required = ['Room Number', 'Number of Rows', 'Number of Bench', 'Number of Student per Bench', 'Left Name', 'Middle Name', 'Right Name']
            if not all(col in df.columns for col in required):
                messagebox.showerror("Error", f"Missing columns: {', '.join(required)}")
                return
            self.room_details_df = df
            self.students_per_bench = int(df['Number of Student per Bench'].iloc[0])
            self.roll_number_indices = [0] * self.students_per_bench
            self.status_label.config(text="✅ Room details loaded. Now select roll number files.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read Excel file: {e}")

    def load_roll_file(self, position):
        filepath = filedialog.askopenfilename(title=f"Select {position} Roll Numbers File", filetypes=[("Excel Files", "*.xlsx *.xls")])
        if not filepath:
            return
        try:
            df = pd.read_excel(filepath)
            if 'Roll Number' not in df.columns:
                messagebox.showerror("Error", f"'Roll Number' column missing in {position} file.")
                return
            self.roll_paths[position] = df['Roll Number'].dropna().tolist()
            self.roll_files_selected[position] = True
            if all(self.roll_files_selected[pos] for pos in ["Left", "Middle", "Right"][:self.students_per_bench]):
                self.roll_numbers_lists = [self.roll_paths[pos] for pos in ["Left", "Middle", "Right"][:self.students_per_bench]]
                self.generate_button.config(state=tk.NORMAL)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read {position} file: {e}")

    def generate_chart(self):
        try:
            wb = openpyxl.Workbook()
            for idx, row in self.room_details_df.iterrows():
                room_number = row['Room Number']
                benches = int(row['Number of Bench'])
                rows = int(row['Number of Rows'])
                ws = wb.create_sheet(title=f"Room {idx+1}")

                # Merge title across all columns
                total_cols = benches * (self.students_per_bench + 1)
                merge_end = openpyxl.utils.get_column_letter(total_cols)
                ws.merge_cells(f"A1:{merge_end}1")
                ws['A1'] = f"ROOM {room_number}"
                ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
                ws['A1'].fill = PatternFill(start_color="3f51b5", end_color="3f51b5", fill_type="solid")
                ws['A1'].alignment = Alignment(horizontal='center')

                # Blank row after title
                ws.append([''])

                # Optional: add student names above each seat column
                left_name = row.get('Left Name', '')
                middle_name = row.get('Middle Name', '')
                right_name = row.get('Right Name', '')
                col_name = 1
                if left_name or middle_name or right_name:
                    for b in range(benches):
                        ws.cell(row=2, column=col_name, value=left_name).alignment = Alignment(horizontal='center')
                        ws.cell(row=2, column=col_name).font = Font(bold=True)
                        ws.cell(row=2, column=col_name+1, value=middle_name).alignment = Alignment(horizontal='center')
                        ws.cell(row=2, column=col_name+1).font = Font(bold=True)
                        ws.cell(row=2, column=col_name+2, value=right_name).alignment = Alignment(horizontal='center')
                        ws.cell(row=2, column=col_name+2).font = Font(bold=True)
                        if b < benches - 1:
                            col_name += self.students_per_bench + 1
                        else:
                            col_name += self.students_per_bench

                # Add row headers (Row 1, Row 2, ...)
                col = 1
                for b in range(benches):
                    ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + self.students_per_bench - 1)
                    cell = ws.cell(row=3, column=col, value=f"Row {b+1}")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="7e57c2", end_color="7e57c2", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                    if b < benches - 1:
                        col += self.students_per_bench + 1
                    else:
                        col += self.students_per_bench

                # Add seat labels (F-1, S-1, T-1)
                seat_label_row = 4
                col = 1
                for b in range(benches):
                    for p in range(self.students_per_bench):
                        cell = ws.cell(row=seat_label_row, column=col + p, value=SEAT_POSITIONS[p])
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = HEADER_FILL
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = BORDER
                    if b < benches - 1:
                        col += self.students_per_bench + 1
                    else:
                        col += self.students_per_bench

                # Fill student roll numbers vertically below each seat
                data_start_row = 5
                indices = [0] * self.students_per_bench
                col = 1
                for b in range(benches):
                    for p in range(self.students_per_bench):
                        for r in range(rows):
                            data_list = self.roll_numbers_lists[p]
                            idx_val = indices[p]
                            value = data_list[idx_val] if idx_val < len(data_list) else ""
                            cell = ws.cell(row=data_start_row + r, column=col + p, value=value)
                            cell.alignment = Alignment(horizontal='center')
                            cell.border = BORDER
                            if (r + b) % 2 == 0:
                                cell.fill = PatternFill(start_color="e0f7fa", end_color="e0f7fa", fill_type="solid")
                            else:
                                cell.fill = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
                            indices[p] += 1
                    if b < benches - 1:
                        col += self.students_per_bench + 1
                    else:
                        col += self.students_per_bench

                self.progress_var.set((idx + 1) / len(self.room_details_df) * 100)
                self.master.update_idletasks()

            del wb['Sheet']
            output_path = os.path.join(os.getcwd(), "SeatingChart_Output.xlsx")
            wb.save(output_path)
            self.generated_file_path = output_path
            self.download_label.config(text=f"✔️ File generated: {output_path}")
            self.download_button.config(state=tk.NORMAL)
            self.open_button.config(state=tk.NORMAL)
            messagebox.showinfo("Success", f"Seating chart saved to:\n{output_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate seating chart:\n{e}")

    def download_file(self):
        if self.generated_file_path:
            save_path = filedialog.asksaveasfilename(title="Save Seating Chart As", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if save_path:
                with open(self.generated_file_path, 'rb') as src:
                    with open(save_path, 'wb') as dst:
                        dst.write(src.read())
                messagebox.showinfo("Downloaded", f"File saved to:\n{save_path}")

    def open_file(self):
        if self.generated_file_path and os.path.exists(self.generated_file_path):
            try:
                if os.name == 'nt':
                    os.startfile(self.generated_file_path)
                elif os.name == 'posix':
                    subprocess.call(['open', self.generated_file_path])
            except Exception as e:
                messagebox.showerror("Error", f"Unable to open file: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = SeatingChartApp(root)
    root.mainloop()
