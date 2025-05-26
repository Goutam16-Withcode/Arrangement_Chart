import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import os
import subprocess

SEAT_POSITIONS = ["F-1", "S-1", "T-1"]
HEADER_FILL = PatternFill(start_color="D1C4E9", end_color="D1C4E9", fill_type="solid")
BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

class SeatingChartApp:
    def __init__(self, master):
        self.master = master
        master.title("Seating and Attendance Sheet Generator")
        master.geometry("600x600")
        master.configure(bg="#f0f0f0")
        self.room_details_df = None
        self.students_per_bench = None
        self.roll_numbers_lists = []
        self.roll_paths = {}
        self.roll_files_selected = {}
        self.generated_file_path = None

        # Set up color scheme
        self.button_bg = "#FFB300"      # Amber tone for buttons
        self.button_hover = "#FFA000"   # Darker amber for hover
        self.header_bg = "#FFB300"      # Amber for headers
        self.frame_bg = "#FFFFFF"       # White for card backgrounds

        # Main title
        title_label = tk.Label(master, text="Seating Chart & Attendance Generator",
                               bg=self.header_bg, fg="white",
                               font=("Helvetica", 18, "bold"))
        title_label.pack(fill="x", pady=(10, 10))

        # Input files frame (card style)
        frame_input = tk.Frame(master, bg=self.frame_bg, bd=2, relief=tk.RIDGE)
        frame_input.pack(padx=10, pady=5, fill="x")
        input_header = tk.Label(frame_input, text="📂 Input Files",
                                bg=self.header_bg, fg="white",
                                font=("Helvetica", 16, "bold"))
        input_header.pack(fill="x")

        # Room Layout upload section
        lbl_room = tk.Label(frame_input, text="Room Layout File:",
                            font=("Helvetica", 12, "bold"), bg=self.frame_bg)
        lbl_room.pack(anchor="w", pady=(10, 0), padx=10)
        self.create_button(frame_input, "📁 Upload Room Layout File", self.load_room_file)

        # Roll Numbers upload section
        lbl_roll = tk.Label(frame_input, text="Roll Numbers Files:",
                            font=("Helvetica", 12, "bold"), bg=self.frame_bg)
        lbl_roll.pack(anchor="w", pady=(10, 0), padx=10)
        for pos in ["Left", "Middle", "Right"]:
            self.roll_files_selected[pos] = False
            self.create_button(frame_input, f"🧾 Upload {pos} Roll Numbers", lambda p=pos: self.load_roll_file(p))

        # Actions frame (card style)
        frame_actions = tk.Frame(master, bg=self.frame_bg, bd=2, relief=tk.RIDGE)
        frame_actions.pack(padx=10, pady=10, fill="x")
        actions_header = tk.Label(frame_actions, text="⚙️ Actions",
                                  bg=self.header_bg, fg="white",
                                  font=("Helvetica", 16, "bold"))
        actions_header.pack(fill="x")

        # Generate, Save, Open buttons
        self.generate_button = self.create_button(frame_actions, "✅ Generate Seating & Attendance",
                                                 self.generate_chart, active=False)
        self.download_button = self.create_button(frame_actions, "⬇️ Save File As...",
                                                 self.download_file, active=False)
        self.open_button = self.create_button(frame_actions, "📂 Open File",
                                              self.open_file, active=False)

        # Progress bar with rounded style
        self.progress_var = tk.DoubleVar()
        style = ttk.Style(master)
        style.theme_use('clam')
        style.configure("TProgressbar", thickness=20, troughcolor="#EEEEEE", background=self.button_bg)
        self.progress = ttk.Progressbar(frame_actions, orient="horizontal", length=400,
                                        mode="determinate", variable=self.progress_var, style="TProgressbar")
        self.progress.pack(pady=20)

        # Status label for messages
        self.status_label = tk.Label(frame_actions, text="", fg="green",
                                     bg=self.frame_bg, font=("Helvetica", 10))
        self.status_label.pack()

    def create_button(self, parent, text, command, active=True):
        """Helper to create a styled button with hover effect."""
        btn = tk.Button(parent, text=text, command=command, width=30, pady=5,
                        bg=self.button_bg, fg="white",
                        font=("Helvetica", 12, "bold"),
                        relief=tk.FLAT, activebackground=self.button_hover, cursor="hand2")
        if not active:
            btn.config(state=tk.DISABLED)
        btn.pack(pady=5)
        if active:
            btn.bind("<Enter>", lambda e: btn.config(bg=self.button_hover))
            btn.bind("<Leave>", lambda e: btn.config(bg=self.button_bg))
        return btn

    def load_room_file(self):
        filepath = filedialog.askopenfilename(title="Select Room Layout Excel File",
                                              filetypes=[("Excel Files", "*.xlsx *.xls")])
        if not filepath:
            return
        try:
            df = pd.read_excel(filepath)
            df.columns = df.columns.str.strip()
            required = ['Room Number', 'Number of Rows', 'Number of Bench',
                        'Number of Student per Bench', 'Left Name', 'Middle Name', 'Right Name']
            if not all(col in df.columns for col in required):
                messagebox.showerror("Error", "Room layout file is missing required columns.")
                return
            self.room_details_df = df
            self.students_per_bench = int(df['Number of Student per Bench'].iloc[0])
            self.status_label.config(text="Room layout loaded. Upload roll number files.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read room layout: {e}")

    def load_roll_file(self, position):
        filepath = filedialog.askopenfilename(title=f"Select {position} Roll Numbers File",
                                              filetypes=[("Excel Files", "*.xlsx *.xls")])
        if not filepath:
            return
        try:
            df = pd.read_excel(filepath)
            if 'Roll Number' not in df.columns:
                messagebox.showerror("Error", f"'{position}' file missing 'Roll Number' column.")
                return
            self.roll_paths[position] = df['Roll Number'].dropna().tolist()
            self.roll_files_selected[position] = True
            self.status_label.config(text=f"{position} roll numbers loaded.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read {position} roll numbers: {e}")

        # Enable generate button if all necessary roll files are loaded
        if self.students_per_bench:
            needed = ["Left", "Middle", "Right"][:self.students_per_bench]
        else:
            needed = []
        if all(self.roll_files_selected.get(pos, False) for pos in needed):
            # Order lists by seat position
            self.roll_numbers_lists = [self.roll_paths[pos] for pos in ["Left", "Middle", "Right"][:self.students_per_bench]]
            self.generate_button.config(state=tk.NORMAL)
            self.generate_button.config(bg=self.button_bg)  # Ensure button shows active color

    def generate_chart(self):
        try:
            wb = openpyxl.Workbook()
            for idx, row in self.room_details_df.iterrows():
                room_number = row['Room Number']
                benches = int(row['Number of Bench'])
                rows = int(row['Number of Rows'])
                ws = wb.create_sheet(title=f"Room {room_number}")

                # Merge title across columns and center
                total_cols = benches * (self.students_per_bench + 1)
                merge_end = openpyxl.utils.get_column_letter(total_cols)
                ws.merge_cells(f"A1:{merge_end}1")
                ws['A1'] = f"ROOM {room_number}"
                ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
                ws['A1'].fill = PatternFill(start_color="3f51b5", end_color="3f51b5", fill_type="solid")
                ws['A1'].alignment = Alignment(horizontal='center')
                ws.append([''])  # blank row

                # Seat position names (Left, Middle, Right) if provided
                left_name = row.get('Left Name', '')
                middle_name = row.get('Middle Name', '')
                right_name = row.get('Right Name', '')
                col_name = 1
                for b in range(benches):
                    if self.students_per_bench >= 1:
                        cell = ws.cell(row=2, column=col_name, value=left_name)
                        cell.alignment = Alignment(horizontal='center')
                        cell.font = Font(bold=True)
                    if self.students_per_bench >= 2:
                        cell = ws.cell(row=2, column=col_name+1, value=middle_name)
                        cell.alignment = Alignment(horizontal='center')
                        cell.font = Font(bold=True)
                    if self.students_per_bench >= 3:
                        cell = ws.cell(row=2, column=col_name+2, value=right_name)
                        cell.alignment = Alignment(horizontal='center')
                        cell.font = Font(bold=True)
                    if b < benches - 1:
                        col_name += self.students_per_bench + 1
                    else:
                        col_name += self.students_per_bench

                # Row headers (Row 1, Row 2, ...)
                col = 1
                for b in range(benches):
                    ws.merge_cells(start_row=3, start_column=col,
                                   end_row=3, end_column=col + self.students_per_bench - 1)
                    cell = ws.cell(row=3, column=col, value=f"Row {b+1}")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="7e57c2", end_color="7e57c2", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                    if b < benches - 1:
                        col += self.students_per_bench + 1
                    else:
                        col += self.students_per_bench

                # Seat labels (F-1, S-1, T-1)
                seat_label_row = 4
                col = 1
                for b in range(benches):
                    for p in range(self.students_per_bench):
                        cell = ws.cell(row=seat_label_row, column=col + p, value=SEAT_POSITIONS[p])
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = HEADER_FILL
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = BORDER
                    if b < benches - 1:
                        col += self.students_per_bench + 1
                    else:
                        col += self.students_per_bench

                # Fill student roll numbers vertically under each seat column
                data_start_row = 5
                indices = [0] * self.students_per_bench
                col = 1
                for b in range(benches):
                    for p in range(self.students_per_bench):
                        for r in range(rows):
                            data_list = self.roll_numbers_lists[p]
                            idx_val = indices[p]
                            value = data_list[idx_val] if idx_val < len(data_list) else ""
                            cell = ws.cell(row=data_start_row + r, column=col + p, value=value)
                            cell.alignment = Alignment(horizontal='center')
                            cell.border = BORDER
                            # Alternate row shading
                            if (r + b) % 2 == 0:
                                cell.fill = PatternFill(start_color="e0f7fa", end_color="e0f7fa", fill_type="solid")
                            else:
                                cell.fill = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
                            indices[p] += 1
                    if b < benches - 1:
                        col += self.students_per_bench + 1
                    else:
                        col += self.students_per_bench

                # Update progress bar
                self.progress_var.set((idx + 1) / len(self.room_details_df) * 100)
                self.master.update_idletasks()

                # Create Attendance worksheet for this room
                att_ws = wb.create_sheet(title=f"Attendance - Room {room_number}")
                blocks = self.students_per_bench
                total_cols_att = blocks * self.students_per_bench + (blocks - 1)
                merge_end_att = openpyxl.utils.get_column_letter(total_cols_att)
                att_ws.merge_cells(f"A1:{merge_end_att}1")
                att_ws['A1'] = f"Attendance - ROOM {room_number}"
                att_ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
                att_ws['A1'].fill = PatternFill(start_color="3f51b5", end_color="3f51b5", fill_type="solid")
                att_ws['A1'].alignment = Alignment(horizontal='center')

                # Section headers (F-1, S-1, T-1)
                section_row = 2
                col_att = 1
                for p, label in enumerate(SEAT_POSITIONS[:self.students_per_bench]):
                    att_ws.merge_cells(start_row=section_row, start_column=col_att,
                                       end_row=section_row, end_column=col_att + self.students_per_bench - 1)
                    cell = att_ws.cell(row=section_row, column=col_att, value=label)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="7e57c2", end_color="7e57c2", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                    col_att += self.students_per_bench + 1

                # Attendance table headers
                header_row = 3
                col_att = 1
                for p in range(self.students_per_bench):
                    headers = ["Serial Number", "Student Roll Number", "Signature"]
                    cell = att_ws.cell(row=header_row, column=col_att, value=headers[0])
                    cell.font = Font(bold=True)
                    cell.fill = HEADER_FILL
                    cell.alignment = Alignment(horizontal='center')
                    att_ws.cell(row=header_row, column=col_att+1, value=headers[1]).font = Font(bold=True)
                    att_ws.cell(row=header_row, column=col_att+1).fill = HEADER_FILL
                    att_ws.cell(row=header_row, column=col_att+1).alignment = Alignment(horizontal='center')
                    att_ws.cell(row=header_row, column=col_att+2, value=headers[2]).font = Font(bold=True)
                    att_ws.cell(row=header_row, column=col_att+2).fill = HEADER_FILL
                    att_ws.cell(row=header_row, column=col_att+2).alignment = Alignment(horizontal='center')
                    col_att += self.students_per_bench + 1

                # Fill attendance data (only students actually seated)
                start_data_row = 4
                col_att = 1
                for p in range(self.students_per_bench):
                    total_seats = benches * rows
                    rolls = self.roll_numbers_lists[p]
                    count = min(len(rolls), total_seats)
                    for j in range(count):
                        att_ws.cell(row=start_data_row+j, column=col_att, value=j+1).alignment = Alignment(horizontal='center')
                        att_ws.cell(row=start_data_row+j, column=col_att+1, value=rolls[j]).alignment = Alignment(horizontal='center')
                        att_ws.cell(row=start_data_row+j, column=col_att+2, value="").alignment = Alignment(horizontal='center')
                    col_att += self.students_per_bench + 1

            # Remove default sheet and save output
            del wb['Sheet']
            output_path = os.path.join(os.getcwd(), "SeatingChart_Output.xlsx")
            wb.save(output_path)
            self.generated_file_path = output_path
            self.download_button.config(state=tk.NORMAL)
            self.open_button.config(state=tk.NORMAL)
            messagebox.showinfo("Success", f"Seating chart and attendance saved to:\n{output_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate charts: {e}")

    def download_file(self):
        if self.generated_file_path:
            save_path = filedialog.asksaveasfilename(title="Save As", defaultextension=".xlsx",
                                                     filetypes=[("Excel Files", "*.xlsx")])
            if save_path:
                with open(self.generated_file_path, 'rb') as src:
                    with open(save_path, 'wb') as dst:
                        dst.write(src.read())
                messagebox.showinfo("Saved", f"File saved to:\n{save_path}")

    def open_file(self):
        if self.generated_file_path and os.path.exists(self.generated_file_path):
            try:
                if os.name == 'nt':
                    os.startfile(self.generated_file_path)
                elif os.name == 'posix':
                    subprocess.call(['open', self.generated_file_path])
            except Exception as e:
                messagebox.showerror("Error", f"Unable to open file: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = SeatingChartApp(root)
    root.mainloop()